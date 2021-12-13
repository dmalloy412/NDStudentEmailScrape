from selenium import webdriver
from selenium.webdriver import ActionChains
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
#import pandas as pd
import openpyxl

#LOGIN TO FSU OUTLOOK
driver = webdriver.Firefox(executable_path=r'C:\Users\Danny Malloy\Documents\DORMBROKER PYTHON\geckodriver.exe')
driver.get("https://login.live.com/login.srf?wa=wsignin1.0&rpsnv=13&ct=1625930197&rver=7.0.6737.0&wp=MBI_SSL&wreply=https%3a%2f%2foutlook.live.com%2fowa%2f%3fnlp%3d1%26RpsCsrfState%3d2fcae0dd-393d-e34c-01cf-afa83000127b&id=292841&aadredir=1&CBCXT=out&lw=1&fl=dob%2cflname%2cwld&cobrandid=90015")
driver.maximize_window()

# email = "danny.malloy@sagenceconsulting.com"
email = "sbc19b@my.fsu.edu"
# password = "Wumboman1230!"
password = "Valencia0711!"

WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//input[@type='email']")))
email_input = driver.find_element_by_xpath("//input[@type='email']")
email_input.send_keys(email)

WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//input[@type='submit']")))
time.sleep(1)
nextbutton = driver.find_element_by_xpath("//input[@type='submit']")
ActionChains(driver).click(nextbutton).perform()

time.sleep(3)
WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//input[@type='password']")))
email_input = driver.find_element_by_xpath("//input[@type='password']")
email_input.send_keys(password)

WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//input[@type='submit']")))
time.sleep(1)
nextbutton = driver.find_element_by_xpath("//input[@type='submit']")
ActionChains(driver).click(nextbutton).perform()

#VERIFICATION

#Contact Collection
WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//div//span[text()='Outlook']")))
contacts = driver.find_element_by_xpath("//a[@title='People']")
ActionChains(driver).click(contacts).perform()

WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//button[@title='Default Global Address List']")))
global_contacts = driver.find_element_by_xpath("//button[@title='Default Global Address List']")
driver.execute_script("arguments[0].scrollIntoView();", global_contacts)
ActionChains(driver).click(global_contacts).perform()

time.sleep(10)
WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//div[@aria-label='Item list' and contains(@class,'ReactVirtualized__Grid')]")))
list_item = driver.find_element_by_xpath("//div[@aria-label='Item list' and contains(@class,'ReactVirtualized__Grid')]")

time.sleep(1)
names_and_emails = {}


# repeat_counter = 0
for x in range(0,2500):
#     print(x)
#     len_old = len(names_and_emails)
#     raw_mail = driver.find_elements_by_xpath("//div[@aria-label='Item list' and contains(@class,'ReactVirtualized__Grid')]//div[@role='listitem']//div[@class='_1e575ojPJe-V4KEJevuytV' or @class='_3T5mthtB7fjE3yLUFJ4QPN']")
#     current_emails = []
#     for email in raw_mail:
#         current_emails.append(email.text)
#     raw_names = driver.find_elements_by_xpath("//div[@aria-label='Item list' and contains(@class,'ReactVirtualized__Grid')]//div[@role='listitem']//div[@class='_2DX0iPG8PDF3Si_o5PlzIj']")
#     current_names = []
#     for i,name in enumerate(raw_names):
#         current_names.append(name.text)
#     names_and_emails.update(dict(zip(current_names, current_emails)))
#     driver.execute_script("arguments[0].scrollTo(0,60*18*{})".format(x), list_item)
#     if (x % 10) == 0:
#         print(names_and_emails, len(names_and_emails))
#     len_new = len(names_and_emails)
#     # if len_new == len_old:
#     #     repeat_counter += 1
#     #     if repeat_counter > 4:
#     #         break
#     flag = True
#     while flag == True:
#         #print("IM IN IT!")
#         try:
#             WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//div[@class='ReactVirtualized__Grid__innerScrollContainer' and @role='group' and not(contains(@style,'pointer-events: none;'))]")))
#             #element_test = driver.find_element_by_xpath("//div[@class='ReactVirtualized__Grid__innerScrollContainer' and @role='group' and not(contains(@style,'pointer-events: none;'))]")
#             #h_of_first_new = int(driver.find_element_by_xpath("(//div[@class='ReactVirtualized__Grid__innerScrollContainer' and @role='group' and not(contains(@style,'max-height ...'))]//div[@class='' and contains(@style,'height: 60px')])[1]").get_attribute("style").split("top: ",1)[1].split("px;",1)[0])
#         except:
#             time.sleep(1)
#             #print("sleep")
#         else:
#             #print("worked")
#             flag = False
#             time.sleep(2)

workbook = openpyxl.Workbook()
sheet = workbook.active

row = 1
for name,email in names_and_emails.items():
    if '@my.fsu.edu' in email:
    #if '@sagenceconsulting.com' in email:
        sheet.cell(row=row, column=1, value=name)
        sheet.cell(row=row, column=2, value=email)
        row += 1

workbook.save(filename="fsu_emails_add2.xlsx")
#workbook.save(filename="sagence_emails2.xlsx")
